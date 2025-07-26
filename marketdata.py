"""
This file contains the functions to retrieve the data from the KuCoin API.
It also contains the functions to write the data to an Excel file and to represent the data in a candlestick chart.
    Author: Nessico (MrSh4d0w)

"""

from datetime import datetime
from kucoin.client import Market
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from openpyxl.styles import Alignment
import threading
import math 
from openpyxl.utils import get_column_letter


#  MarketData

"""
    Client to obtain data, we access to kucoin api
"""
client = Market(url='https://api.kucoin.com')
client_futures = Market(url='https://api-futures.kucoin.com')

"""
    TIMESTAMP CONVERSION

"""
def getServerTimestamp():
    """Get the server timestamp.

    Returns:
        timestamp(int): timestamp in milliseconds
    """
    return client.get_server_timestamp()

def getServerTime():
    """Get the server time.

    Returns:
        datetime: server time
    """
    return datetime.fromtimestamp(getServerTimestamp()/1000)

def getDatetime(year, month, day, hour, minute, second):
    """Define a date in datetime format.

    Args:
        year (int): year
        month (int): month
        day (int): day
        hour (int): hour
        minute (int): minute
        second (int): second

    Returns:
        datetime: date
    """
    return datetime(year, month, day, hour, minute, second)

def getTimestamp_to_date(timestamp):
    """Convert timestamp (ms) to date (day, month, year).

    Args:
        timestamp (int): timestamp in milliseconds

    Returns:
        datetime: timestamp in date format (day, month, year)
    """
    return datetime.fromtimestamp(int(timestamp)).strftime('%d-%m-%Y')

def getTimestamp_to_datetime(timestamp):
    """Convert timestamp (ms) to date.

    Returns:
        datetime: timestamp in date format
    """
    return datetime.fromtimestamp(int(timestamp/1000))
    
def getDatetime_to_timestamp(date):
    """Convert date to timestamp (ms).
    It returned in ms to be consistent with the server timestamp.
    KuCoin uses ms for the timestamp.
    Returns:
        int: date in timestamp format (ms)
    """
    return int(date.timestamp())

def getNowTimestamp():
    """Get the current timestamp.

    Returns:
        int: current timestamp in milliseconds
    """
    return getDatetime_to_timestamp(datetime.now())

def getNowDate():
    """Get the current date.

    Returns:
        datetime: current date
    """
    return datetime.now()

"""
    DATA PROCESSING

"""


"""
    WRITE ON EXCEL THE DATA RETRIEVED
"""


def write_headerpage(ws, coin, interval, start=0, end=0, lastrow=0):	
    ws.append(["coin", "interval", "start", "end", "lastrow"])
    ws.append([f"{coin}", f"{interval}", f"{start}", f"{end}", f"{lastrow}"])
    ws.append(["", "", "", "", "", "", ""])
    return ws

def writeDataToExcel(ws, data, filename):
    timestamps, open_prices, close_prices, high_prices, low_prices, volumes = ([] for _ in range(6))

    # Iterate over the data and write them to the Excel file
    for sublist in data:
        timestamp, open_price, close_price, high_price, low_price, volume = sublist

        # Add the values to the current row
        row = [timestamp, float(open_price), float(close_price), float(high_price), float(low_price), float(volume)]
        ws.append(row)

        # Append the values to the lists
        timestamps.append(timestamp)
        open_prices.append(float(open_price))
        close_prices.append(float(close_price))
        high_prices.append(float(high_price))
        low_prices.append(float(low_price))
        volumes.append(float(volume))


def write_to_excel(data, filename, withHeader=True):
    """Write the data to an Excel file.

    Args:
        data (list): list of lists containing the data
        filename (str): name of the file
        withHeader (bool): decide if you want a header on excel to put the last value, the coin, startDate... Default: True
    """
    # Create a new workbook
    wb = Workbook()

    # Select the active sheet
    ws = wb.active

    if withHeader:
        # Tokenizar el string usando el punto como separador
        tokens = filename.split('.')

        # Asignar cada parte a su respectiva variable
        coin = tokens[0]
        interval = tokens[1]
        start = tokens[2]
        end = tokens[3]
        
        ws.append(["coin", "interval", "start", "end", "lastrow"])
        ws.append([f"{coin}", f"{interval}", f"{start}", f"{end}"])
        ws.append(["", "", "", "", "", "", ""])

    # Add headers
    ws.append(["Timestamp", "Open", "Close", "High", "Low", "Volume"])

    # Create empty lists to store the values
    timestamps, open_prices, close_prices, high_prices, low_prices, volumes = ([] for _ in range(6))

    lastrow = 4

    # Iterate over the data and write them to the Excel file
    for sublist in data:
        timestamp, open_price, close_price, high_price, low_price, volume = sublist

        # Add the values to the current row
        row = [timestamp, float(open_price), float(close_price), float(high_price), float(low_price), float(volume)]
        ws.append(row)
        lastrow += 1

        # Append the values to the lists
        timestamps.append(timestamp)
        open_prices.append(float(open_price))
        close_prices.append(float(close_price))
        high_prices.append(float(high_price))
        low_prices.append(float(low_price))
        volumes.append(float(volume))

    ws['E2'] = lastrow

    # Save the Excel file
    wb.save(filename)

    print(f"\nData has been written to {filename}")


def write_dataframe_to_excel(dataframe, excel_file):
    # Header
    header = ["Timestamp", "Open", "Close", "High", "Low", "Volume", "base_volume"]
    
    # Extract additional information from the filename
    tokens = excel_file.split('.')
    
    # Ensure there are enough tokens to avoid errors
    if len(tokens) < 4:
        raise ValueError("The filename must be in the format 'coin.interval.start.end.xlsx'")
    
    coin, interval, start, end = tokens[0], tokens[1], tokens[2], tokens[3]
    
    # Create a workbook and worksheet
    workbook = Workbook()
    ws = workbook.active
    
    # Write the additional information and the header
    ws.append(["coin", "interval", "start", "end", "lastrow"])
    ws.append([coin, interval, start, end, len(dataframe)+4])
    ws.append([""] * 5)  # Empty row to separate additional information from the DataFrame
    
    ws.append(header)
    
    # Ensure the DataFrame contains the columns in the order of the header
    # Fill any missing columns with zeros if necessary
    for col in header:
        if col not in dataframe.columns:
            dataframe[col] = 0

    
    dataframe = dataframe[header]
    
    # Write the DataFrame rows to the worksheet and adjust the column width
    row_num = 0
    for row in dataframe.itertuples(index=False, name=None):
        if row_num in [1, 2, 4]:
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal='center')
        row_num += 1
        ws.append(row)

    column_widths = 15 
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = column_widths
    
    # Center align the specified rows in the Excel file
    for row_num in [1, 2, 4]:
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center')

    # Save the file
    workbook.save(excel_file)
    print(f"DataFrame written to the file {excel_file}")


"""
    GET DATA FROM KUCOIN API
    
"""
    
def getData(coin, interval, start, end, shared_data, writeOnExcel=False):
    """Get the data from the KuCoin API.
    Due to limitations in KuCoin API it only can retrieve 1500 elements.
    
    Args:
        coin (str): coin pair (e.g. BTC-USDT)
        interval (str): interval (1min, 5min, 15min, 1hour, 4hour, 8hour, 1day, 1week, 1month)
        start (int): start timestamp
        end (int): end timestamp
        writeOnExcel (bool): write the data to an Excel file (default: False)

    Returns:
        list: data
    """
    #print(f"Getting data for {coin} from {start} to {end}")
    data = client.get_kline(coin, interval, startAt=start, endAt=end)
    file = f"{coin}.{interval}.{start}.{end}.xlsx"

    if writeOnExcel:
        write_to_excel(data, file)
    
    shared_data.extend(data)

    return data

"""
    MULTI-THREADING

"""

# Temporalidades en diccionario y su calculo
temporalities = {
        '1min': 1500 * 60,
        '5min': 1500 * 60 * 5,
        '15min': 1500 * 60 * 15,
        '1hour': 1500 * 60 * 60,
        '4hour': 1500 * 60 * 60 * 4,
        '8hour': 1500 * 60 * 60 * 8,
        '1day': 1500 * 60 * 60 * 24,
        '1week': 1500 * 60 * 60 * 24 * 7,
        '1month': 1500 * 60 * 60 * 24 * 30
}


def multi_threading(start_timestamp, end_timestamp, temporality, coin, writeOnExcel=True):
    """Retrieve the data in multiple threads.

    Args:
        start_timestamp (int): start timestamp
        end_timestamp (int): end timestamp
        temporality (str): temporality (1min, 5min, 15min, 1hour, 4hour, 8hour, 1day, 1week, 1month)
        coin (str): coin pair (e.g. BTC-USDT)
        writeOnExcel (bool): decide if you want to write the retrieved data on Excel. Default value: True
    
    Returns:
        list: shared list with the data
        excelFile (str): name of the Excel file where the data is stored
    """

    num_threads = math.ceil((end_timestamp - start_timestamp) / temporalities[temporality])
    threads = []
    shared_data = []
    lock = threading.Lock()

    def thread_safe_getData(coin, temporality, start, end, shared_data, writeOnExcel):
        try:
            data = getData(coin, temporality, start, end, [], writeOnExcel)
            with lock:
                shared_data.extend(data)
        except Exception as e:
            print(f"Error en hilo para {coin} ({start} - {end}): {e}")


    for i in range(num_threads):
        # We control the ranges with the i variable
        # print(f"Thread {i+1} of {num_threads}")
        # print("Start timestamp:", start_timestamp + i * temporalities[temporality])
        
        # Calculate the range of timestamps for the current thread
        start = start_timestamp + i * temporalities[temporality]

        # Calculate the end timestamp for the current thread
        end = min(start_timestamp + (i + 1) * temporalities[temporality], end_timestamp)

        thread = threading.Thread(target=thread_safe_getData, args=(coin, temporality, start, end, shared_data, False))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    # Sort the shared data by timestamp (the first attribute) to ensure it is in order
    shared_data.sort(key=lambda x: x[0])

    if writeOnExcel:
        # Write the data to an Excel file
        file = f'{coin}.{temporality}.{getTimestamp_to_date(start_timestamp)}.{getTimestamp_to_date(end_timestamp)}'
        excelFile = file + '.xlsx'

        # Write the data to an Excel file
        header = ["Timestamp", "Open", "Close", "High", "Low", "Volume", "base_volume"]
        df = pd.DataFrame(shared_data, columns=header)

        write_dataframe_to_excel(df, excelFile)

        return df, excelFile
    else:
        header = ["Timestamp", "Open", "Close", "High", "Low", "Volume", "base_volume"]
        df = pd.DataFrame(shared_data, columns=header)
        return df, None

"""
    INDICATORS

"""

def getHeader(excelFile):
    """Obtain the header of the Excel file.

    Args:
        excelFile (str): name of the Excel file

    Returns:
        list: header of the Excel file
    """
    wb = load_workbook(excelFile)
    ws = wb.active
    header = []

    for cell in ws[4]:
        header.append(cell.value)

    return header


def addEMAs(mav, excelFile):
    """ Add EMAs to the values of an excel file

    Args:
        mav (array): values of the desired EMAs
        excelFile (string): the name of the excel file (ex: btcusdt.xlsx)
    """
    # The higher ema
    maxEma = max(mav)

    #print(f"The max ema is {maxEma}")

    header = getHeader(excelFile)

    # We update the header
    for m in mav:
        header.append(f"EMA_Close_{m}")
    
    # We load other time the excel
    wb = load_workbook(excelFile)
    ws = wb.active

    # Lee el archivo Excel desde la fila 4 (índice 3 en pandas)
    df_excel = pd.read_excel(excelFile, skiprows=4, header=None)

    # Asigna las columnas a df_excel
    df_excel.columns = header[:len(df_excel.columns)]

    # Llena los valores faltantes con 0
    df_excel = df_excel.reindex(columns=header, fill_value=0)

    # Obtain the data using multi_threading
    data = calculateGapData(excelFile, maxEma)

    # Concatenate the data DataFrame at the beginning of df_excel
    df_excel = pd.concat([data, df_excel], ignore_index=True)
    
    # Add the columns to the dataframe and fill them with 0
    for m in mav:
        df_excel[f'EMA_Close_{m}'] = 0
    
    # Associate the columns with header
    df_excel.columns = header

    # We replace missing values with 0
    df_excel.fillna(0, inplace=True)
    
    # Then we calculate the emas
    for m in mav:
        df_excel[f'EMA_Close_{m}'] = df_excel['Close'].ewm(span=m, adjust=False, min_periods=m).mean()
        
    # Drop the missing values
    df_excel = df_excel.dropna()

    print(excelFile)
    writer = pd.ExcelWriter(excelFile, engine='openpyxl')
    writer._book = wb
    writer._sheets = {ws.title: ws for ws in wb.worksheets}

    # Write it on excel
    df_excel.to_excel(writer, sheet_name='Sheet', startrow=3, index=False, header=True)

    # Adjust the column width
    column_widths = 15 
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = column_widths

    wb.save(excelFile)
    writer.close()


def calculateGapData(excelFile, numValuesRequired):
    """
        Calculate the gap data to retrieve the missing data (numValuesRequired) in the excel file
    
    Args:
        excelFile (string): the name of the excel file (ex: btcusdt.xlsx)
        numValuesRequired (int): the number of values required to calculate the missing data
        
    Returns:
        data (dataFrame): the data of the coin
    """
    wb = load_workbook(excelFile)
    ws = wb.active
    t1 = ws['A5'].value
    t2 = ws['A6'].value
    
    t3 = int(t2) - int(t1)
    
    firstTimestamp = int(t1) - (t3 * numValuesRequired)
    
    data = multi_threading(int(firstTimestamp), int(t1)-int(t3), ws['B2'].value, ws['A2'].value, writeOnExcel=False)[0]

    wb.close()
    
    return data
    
